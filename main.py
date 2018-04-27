from openpyxl import Workbook, load_workbook
from openpyxl.utils import column_index_from_string
from argparse import ArgumentParser

#Probably unneccesary, but exact table from picture
def generate_test():
    wb = Workbook()
    ws = wb.active
    for i in xrange(12):
        for j in xrange(9):
            if j == 0:
                ws.cell(row=i + 1, column=j + 1, value="Row " + str(i + 1))
            elif j > 0 and j < 5:
                ws.cell(row=i + 1, column=j + 1, value="Data " + str(j))
            elif j == 5:
                if i == 0 or i == 1 or i == 3 or i == 5 or i == 8:
                    ws.cell(row=i + 1, column=j + 1, value="Col 1")
            elif j == 6:
                if i > 0 and i < 6:
                    ws.cell(row=i + 1, column=j + 1, value="Col 2")
            elif j == 7:
                if i < 9:
                    ws.cell(row=i + 1, column=j + 1, value="Col 3")
            elif j == 8:
                if (i > 0 and i < 4) or (i > 4 and i < 8) or i == 11:
                    ws.cell(row=i + 1, column=j + 1, value="Col 4")
    wb.save("test2.xlsx")

def main():

    parser = ArgumentParser(description="Sorts Excel Sheets")
    parser.add_argument("-dc", "--datacolumns", help="Amount of data columns", action="store", required=True, type=int)
    parser.add_argument("-c", "--columns", help="Amount of columns to sort", action="store", required=True, type=int)
    parser.add_argument("-f", "--file", help="Input file" action="store", required=True)

    args = parser.parse_args()

    columns = args.columns
    wb = load_workbook(args.file)
    out = Workbook()
    out.remove(out.active)
    get_sheet_to_table(wb, out, columns, args.datacolumns)
    out.save("out.xlsx")

def get_sheet_to_table(wb, out, columns, dcs):
    ws = wb.active
    outcols = {}
    for i in xrange(columns):
        outcols[i] = {'sheet': out.create_sheet(str(i + 1)), 'num': 1, 'rows':[]}
    for i in ws.iter_rows():
        for cell in i:
            if cell.col_idx <= dcs + 1:
                pass
            elif cell.value:
                idx = column_index_from_string(cell.column) - dcs - 2
                outcols[idx]['sheet'].cell(column=dcs + 2, row=outcols[idx]['num'], value = cell.value)
                outcols[idx]['num'] += 1
                outcols[idx]['rows'].append(i)
    for col in outcols.keys():
        outcols[col]['num'] = 1
        for row in outcols[col]['rows']:
            for cell in row:
                if cell.col_idx <= dcs + 1:
                    outcols[col]['sheet'].cell(column=column_index_from_string(cell.column), row=outcols[col]['num'], value=cell.value)
            outcols[col]['num'] += 1


if __name__ == '__main__':
    main()
